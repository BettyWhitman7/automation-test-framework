import time
from collections import deque

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

class ExcelReport:
    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file, data_only=True)
        self.wb.guess_types = True
        self.headers = []
        self.result_map = {"通过": "P", "失败": "F"}

    # 切换到指定名字的sheet页面
    def active_sheet(self, sheet_name):
        self.ws = self.wb[sheet_name]
        self.headers = self.getRowValues(7)

    # 获取所有的sheet name
    def get_sheet_names(self):
        sheetnames = self.wb.sheetnames
        return sheetnames

    def get_value(self, add):
        return self.ws[add].value

    def get_value_by_rc(self, r, c):
        return self.ws.cell(r, c).value

    def set_value(self, add, value):
        if 'H' in add:
            self.ws[add].alignment = Alignment(horizontal='center', vertical='center')
            self.ws[add].border = Border(bottom=Side(style='thin'))
            self.ws[add].font = Font(size=10)
        self.ws[add] = value

    def set_value_by_rc(self, row, cloum, value):

        cell = self.ws.cell(row, cloum).value = value


    # 获取某行所有值
    def getRowValues(self, row):
        columns = min(20, self.ws.max_column)

        rowdata = []
        for i in range(1, columns + 1):
            cellvalue = self.ws.cell(row=row, column=i).value
            rowdata.append(cellvalue)
        return rowdata

    def write_to_excel(self, testcase_name, result, errInfo=None, case_id=None, offset=0):
        """
         # 写入excel里面
        :param testcase_name: 测试用例名称
        :param result: 测试结果
        :param errInfo: 结果错误信息
        :return:
        """
        # 获取最大行数
        for i in range(8, self.ws.max_row + 1):
            # 获取每一行的数据
            row_data = self.getRowValues(i)
            # 如果用例名称一样，那么就在这一行写入结果
            is_case = row_data[self.headers.index('类型')] == "用例"
            if not is_case:
                continue
            is_match = False
            if case_id and row_data[self.headers.index('用例唯一标识符')] == case_id:
                is_match = True
            elif row_data[self.headers.index('用例名称')] == testcase_name:
                is_match = True
            if is_match:
                # 写入结果
                result = self.result_map.get(result, result)
                self.set_value_by_rc(i + offset, self.headers.index('P/F') + 1, result)
                # 写入备注信息
                if errInfo:
                    self.set_value_by_rc(i + offset, self.headers.index('实际结果及其他说明') + 1, errInfo)
                return

    def write_info_by_casename(self, testcase_name, info, cloum_of_casename, cloum_of_res):
        """
         # 根据用例名称往excel里面写入（错误）信息,支持指定列
        :param testcase_name: 测试用例名称
        :param info: 需要写入的信息
        :param cloum_of_casename: 用例名所在列
        :cloum_of_res: 结果所在列
        """
        # 获取最大行数
        for i in range(8, self.ws.max_row + 1):
            # 获取每一行的数据
            row_data = self.getRowValues(i)
            # 如果用例名称一样，那么就在这一行写入结果
            if row_data[cloum_of_casename] == testcase_name:
                # 写入结果
                self.set_value_by_rc(i, cloum_of_res, info)

    def find_line_in_excel(self, cloum, value):
        """
        在 Excel 中查找指定列（cloum）中与给定值（value）匹配的行号。

        遍历从第8行开始到最后一行，每一行的数据通过 getRowValues() 获取，
        当找到该行中指定列的值等于 value 时，即返回该行的行号。

        参数:
            cloum: int
                需要检查的列标，用于确定在哪一列进行匹配。
            value: Any
                要匹配的值，当行中该列的值与此值相等时认为找到了目标行。

        返回:
            int 或 None:
                找到匹配的行号后返回对应的行号，如果没有匹配项则返回 None。
        """
        for i in range(8, self.ws.max_row + 1):
            # 获取每一行的数据
            row_data = self.getRowValues(i)
            # 如果用例名称一样，那么就在这一行写入结果
            if row_data[cloum] == value:
                return i

    def write_value_to_excel(self, pattern, pattern_cloum, value, value_cloum):
        """
        在 Excel 中写入单个值。

        首先在指定的列（pattern_cloum）中查找匹配 pattern 的行，
        然后在该行的指定列（value_cloum）写入给定的 value。

        参数:
            pattern: Any
                用于匹配的模式或关键字，用以在 Excel 中定位目标行。
            pattern_cloum: int
                在 Excel 中查找模式的列索引。
            value: Any
                要写入 Excel 单元格中的值。
            value_cloum: int
                写入值的列索引。
        """
        line_num = self.find_line_in_excel(pattern_cloum, pattern)
        self.set_value_by_rc(line_num, value_cloum, value)

    def write_step_to_excel(self, pattern, pattern_cloum, d: deque, value_cloum):
        """
        在 Excel 中写入多个步骤的结果。

        该函数先根据给定的模式（pattern）在指定的列（pattern_cloum）中找到目标行，
        然后遍历 deque d 中的每个元素，将每个元素中的"结果"写入到目标行下方连续的单元格中，
        同时将"附言"写入相邻的右侧单元格。写入时，"结果"写入到 value_cloum 列，
        "附言"写入到 value_cloum+1 列。

        参数:
            pattern: Any
                用于在 Excel 中定位目标行的模式或关键字。
            pattern_cloum: int
                用于匹配模式的 Excel 列索引。
            d: deque
                一个 deque 对象，包含多个字典，每个字典至少包含以下两个键:
                    "结果": 要写入的结果值。
                    "附言": 要写入的附加说明。
            value_cloum: int
                开始写入的列索引，"结果"写入到此列，"附言"写入到此列的右侧（value_cloum + 1）。
        """
        line_num = self.find_line_in_excel(pattern_cloum, pattern)
        print(f"匹配到该用例在excel第{line_num}行")
        if not line_num:
            raise Exception(f"表中第[{pattern_cloum}]列未找到：{pattern}")
        length_d = len(d)
        print(f"此条用例步骤长度为：{length_d}")
        for i in range(length_d):
            print(f"将在excel第{line_num + d[i]['步骤']}行, 打印结果【{d[i]['结果']}】和附言【{ d[i]['附言']}】")

            self.set_value_by_rc(line_num + d[i]["步骤"], value_cloum, d[i]["结果"])
            self.set_value_by_rc(line_num + d[i]["步骤"], value_cloum + 1, d[i]["附言"])

    def write_info_by_caseID(self, testcase_id, info, cloum_of_info, cloum_of_caseid=0):
        """
         # 根据用例ID往excel里面写入（错误）信息,支持指定列
        :param testcase_id: 测试用例ID
        :param info: 需要写入的信息
        :param cloum_of_info: 结果所在列
        :cloum_of_caseid: 用例ID所在列（默认0）
        """
        # 获取最大行数
        for i in range(8, self.ws.max_row + 1):
            # 获取每一行的数据
            row_data = self.getRowValues(i)
            # 如果用例名称一样，那么就在这一行写入结果
            if row_data[cloum_of_caseid] == testcase_id and row_data[2] == "用例":
                # 写入结果
                self.set_value_by_rc(i, cloum_of_info, info)

    def write_res_by_caseID(self, testcase_id, res, cloum_of_res, cloum_of_caseid=0):
        """
         # 根据ID往excel里面写入用例结果,支持指定列
        :param testcase_id: 测试用例ID
        :param res: 测试结果
        :param cloum_of_res: 结果所在列
        :cloum_of_caseid: 用例ID所在列（默认0）
        """
        # 获取最大行数
        for i in range(8, self.ws.max_row + 1):
            # 获取每一行的数据
            row_data = self.getRowValues(i)
            # print(row_data)
            # 如果用例名称一样，那么就在这一行写入结果
            # print(row_data)
            if str(row_data[cloum_of_caseid]) == testcase_id:
                # 写入结果
                print(row_data)
                print("-----------------------------------------------------"+row_data[cloum_of_caseid])
                print(i, cloum_of_res, res)
                print("-----------------------------------------------------" + row_data[cloum_of_caseid])
                self.set_value_by_rc(i, cloum_of_res, res)

    def save_file(self, add_name):
        self.wb.save(add_name)

    def remove_sheet(self, sheetname):
        """
        删除sheet
        """
        ws = self.wb[sheetname]
        self.wb.remove(ws)

    def do_close(self):
        self.wb.close()

    def do_write_in_excel(self, i, one_result):
        """
        写入到excel表格里面，还需要做单个用例的总结
        :param i:
        :param one_result:
        :return:
        """
        # 该条用例的测试结果

        case_result = None
        if self.get_value("C" + str(i + 1)) == one_result[1]:

            # 写入备注信息
            if one_result[4]:
                self.set_value("H" + str(i + 1 + one_result[2]), one_result[4])
            # 写入结果
            if one_result[3]:
                self.set_value("G" + str(i + 1 + one_result[2]), "P")
            else:
                case_result = False
                self.set_value("G" + str(i + 1 + one_result[2]), "F")
        # 该条用例总结的测试结果
        if case_result:
            self.set_value("G" + str(i + 1), "P")
        elif case_result == False:
            self.set_value("G" + str(i + 1), "F")
        else:
            self.set_value("G" + str(i + 1), case_result)

    def create_sheet(self, sheetname):
        self.ws = self.wb.create_sheet(sheetname)
        return self.ws

    def copy_sheet(self, worksheet):
        self.wb.copy_worksheet(worksheet)

    def get_max_row(self):
        return self.ws.max_row

    def get_max_colum(self):
        return self.ws.max_column


if __name__ == "__main__":
    add1 = r"../reports/fudaiA6_report/福戴A6项目台架全功能测试用例-1221.xlsx"
    # add1 = r"../doc/testcase_dir/福戴A6项目台架全功能测试用例.xlsx"
    myexcel = ExcelReport(add1)
    myexcel.active_sheet('主界面')
    print(myexcel.getRowValues(7))
    d = deque([
        {"步骤": 1,"结果":"通过", "附言":"1"},
        {"步骤": 3,"结果":"不通过", "附言":"3"},
        {"步骤": 2,"结果":"通过", "附言":"2"},
        {"步骤": 4,"结果":"不通过", "附言":"4"}
    ])
    # myexcel.write_info_by_casename("前台FM播放", "pass", 4, 10)
    # myexcel.write_value_to_excel("收音机预置频率", 3, "pass", 10)
    myexcel.write_step_to_excel("主图宫格区滑动", 3, d, 9)
    # myexcel.write_to_excel("嘴部遮挡_1", "通过", "ddsdf")
    # myexcel.write_to_excel("嘴部遮挡_2", "通过", "wgweg")
    myexcel.save_file(add1)
